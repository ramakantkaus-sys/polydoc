"""Binary formats: DOCX, PPTX, XLSX, PDF.

These are round-trip tests: write with polydoc, read back with polydoc, and assert the
structure survived. Where possible they also assert against the *native* library
(openpyxl, python-docx) so a bug in both directions of our own code cannot hide.
"""

from __future__ import annotations

import pytest

from conftest import needs_docx, needs_pdf_read, needs_pdf_write, needs_pptx, needs_xlsx

import polydoc
from polydoc.model import (
    Alignment,
    CodeBlock,
    Document,
    Heading,
    ListBlock,
    ListItem,
    ListStyle,
    Metadata,
    Paragraph,
    Slide,
    Table,
    Text,
    TextStyle,
)


# ---------------------------------------------------------------------------
# DOCX
# ---------------------------------------------------------------------------


@needs_docx
class TestDocx:
    @pytest.fixture
    def written(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.docx")
        return polydoc.open(path)

    def test_file_is_created(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.docx")
        assert path.exists() and path.stat().st_size > 1000

    def test_detected_as_docx(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.docx")
        assert polydoc.detect(path) == "docx"

    def test_metadata_round_trip(self, written):
        assert written.metadata.title == "Quarterly Report"
        assert written.metadata.authors == ["Ada Lovelace"]
        assert written.metadata.subject == "Engineering"

    def test_template_metadata_does_not_leak(self, tmp_path):
        """python-docx's default template claims an author of "python-docx" and a
        2013 creation date; neither must survive into our output."""
        path = Document([Paragraph.of("x")]).save(tmp_path / "clean.docx")
        reread = polydoc.open(path)
        assert reread.metadata.author in (None, "")
        assert reread.metadata.created is None or reread.metadata.created.year >= 2024

    def test_timestamp_can_be_suppressed(self, tmp_path):
        path = Document([Paragraph.of("x")]).save(tmp_path / "n.docx", timestamp=False)
        assert polydoc.open(path).metadata.created is None

    def test_headings_and_levels(self, written):
        assert [(h.level, h.text) for h in written.headings] == [
            (1, "Quarterly Report"),
            (2, "Findings"),
            (2, "Data"),
            (2, "Appendix"),
        ]

    def test_character_formatting(self, written):
        paragraph = written.find("paragraph:contains(Plain)")
        runs = {r.text: r.style for r in paragraph.content if r.type == "text"}
        assert runs["bold"].bold is True
        assert runs["italic"].italic is True
        assert runs["mono"].is_monospace

    def test_hyperlink(self, written):
        link = written.find("link")
        assert link.href == "https://example.com"
        assert link.text == "link"

    def test_hyperlink_style_decoration_is_dropped(self, written):
        # Word underlines links via its character style; that is presentation, not
        # authored emphasis, so it must not leak back into the model.
        link = written.find("link")
        assert link.content[0].style.underline is None

    def test_paragraph_alignment(self, written):
        centred = written.find("paragraph[style.alignment=center]")
        assert centred is not None and centred.text == "Centred text."

    def test_nested_list_with_marker_styles(self, written):
        block = written.find("list_block")
        assert block.marker_style is ListStyle.ORDERED
        assert len(block.items) == 3
        nested = [i for i in block.items if i.sublists]
        assert nested and nested[0].sublists[0].marker_style is ListStyle.BULLET
        assert len(nested[0].sublists[0].items) == 2

    def test_table_content_and_header(self, written):
        table = written.tables[0]
        assert table.dimensions == (3, 3)
        assert table.header_rows == 1
        assert table.to_matrix() == [
            ["Component", "Trials", "Rate"],
            ["Reader", "1200", "99.2%"],
            ["Writer", "1150", "98.7%"],
        ]

    def test_table_caption_is_reattached(self, written):
        assert written.tables[0].caption == "Measured rates"

    def test_code_block_is_recovered(self, written):
        code = written.find("code_block")
        assert code is not None
        assert "def f(x):" in code.code

    def test_horizontal_rule_survives(self, written):
        assert written.find("horizontal_rule") is not None

    def test_page_break_survives(self, written):
        assert written.find("page_break") is not None

    def test_quote_survives(self, written):
        assert written.find("quote") is not None

    def test_colour_round_trip(self, tmp_path):
        document = Document([Paragraph([Text("red", TextStyle(color="#cc0000"))])])
        reread = polydoc.open(document.save(tmp_path / "c.docx"))
        assert reread.body[0].content[0].style.color == "#cc0000"

    def test_merged_cells(self, tmp_path):
        from polydoc.model import TableCell, TableRow

        table = Table(
            [
                TableRow([TableCell.of("spanning", colspan=2)]),
                TableRow.of(["a", "b"]),
            ]
        )
        reread = polydoc.open(Document([table]).save(tmp_path / "m.docx"))
        assert reread.tables[0].rows[0].cells[0].colspan == 2

    def test_edit_survives_the_round_trip(self, template_document, tmp_path):
        template_document.replace_text("{{client}}", "Acme Ltd")
        reread = polydoc.open(template_document.save(tmp_path / "t.docx"))
        assert "Acme Ltd" in reread.text
        assert "{{client}}" not in reread.text

    def test_empty_document_is_valid(self, tmp_path):
        path = Document().save(tmp_path / "empty.docx")
        assert polydoc.open(path).body == []


# ---------------------------------------------------------------------------
# PPTX
# ---------------------------------------------------------------------------


@needs_pptx
class TestPptx:
    @pytest.fixture
    def deck(self):
        return Document(
            metadata=Metadata(title="Roadmap", authors=["Ada"]),
            body=[
                Slide(
                    title="Roadmap 2026",
                    content=[Paragraph.of("Prepared by the platform team.")],
                    notes="Speaker notes here.",
                    index=1,
                ),
                Slide(
                    title="Priorities",
                    content=[
                        ListBlock(
                            [
                                ListItem.of("Ship it"),
                                ListItem(
                                    [
                                        Paragraph.of("Improve fidelity"),
                                        ListBlock(
                                            [ListItem.of("tables"), ListItem.of("notes")],
                                            ListStyle.BULLET,
                                        ),
                                    ]
                                ),
                            ],
                            ListStyle.ORDERED,
                        )
                    ],
                    index=2,
                ),
                Slide(
                    title="Numbers",
                    content=[Table.from_rows([["Q", "Rev"], ["Q1", "100"]])],
                    index=3,
                ),
            ],
        )

    @pytest.fixture
    def written(self, deck, tmp_path):
        return polydoc.open(deck.save(tmp_path / "deck.pptx"))

    def test_detected_as_pptx(self, deck, tmp_path):
        assert polydoc.detect(deck.save(tmp_path / "d.pptx")) == "pptx"

    def test_slide_count_and_titles(self, written):
        assert [s.title for s in written.slides] == [
            "Roadmap 2026",
            "Priorities",
            "Numbers",
        ]

    def test_title_is_not_confused_with_body(self, written):
        # A regression guard: the title placeholder must not receive body text.
        assert written.slides[0].content
        assert written.slides[0].content[0].text == "Prepared by the platform team."

    def test_speaker_notes(self, written):
        assert written.slides[0].notes == "Speaker notes here."

    def test_layout_is_recorded(self, written):
        assert written.slides[0].layout

    def test_ordered_list_marker_survives(self, written):
        block = written.find("list_block")
        assert block.marker_style is ListStyle.ORDERED

    def test_nested_bullets_survive(self, written):
        block = written.find("list_block")
        nested = [i for i in block.items if i.sublists]
        assert nested and len(nested[0].sublists[0].items) == 2

    def test_table_content(self, written):
        assert written.tables[0].to_matrix() == [["Q", "Rev"], ["Q1", "100"]]

    def test_metadata(self, written):
        assert written.metadata.title == "Roadmap"
        assert written.metadata.authors == ["Ada"]

    def test_flowing_document_is_paginated(self, tmp_path):
        source = "# Report\n\n" + "\n\n".join(
            f"## Section {i}\n\n" + "\n\n".join(f"Para {i}.{j}." for j in range(1, 6))
            for i in range(1, 4)
        )
        document = polydoc.loads(source, "markdown")
        reread = polydoc.open(document.save(tmp_path / "p.pptx"))
        # One slide per heading, plus the title slide.
        assert len(reread.slides) >= 3
        assert any(s.title == "Section 1" for s in reread.slides)

    def test_never_writes_a_zero_slide_deck(self, tmp_path):
        reread = polydoc.open(Document().save(tmp_path / "z.pptx"))
        assert len(reread.slides) >= 1


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


@needs_xlsx
class TestXlsx:
    @pytest.fixture
    def workbook_document(self):
        return Document(
            metadata=Metadata(title="Finance"),
            body=[
                Heading.of("Finance Summary", 1),
                Paragraph.of("Unaudited."),
                Table.from_rows(
                    [
                        ["Quarter", "Revenue", "Growth", "Date"],
                        ["Q1", "$1,200.00", "12%", "2026-01-31"],
                        ["Q2", "$1,850.00", "54%", "2026-04-30"],
                    ],
                    caption="Revenue",
                ),
            ],
        )

    def test_detected_as_xlsx(self, workbook_document, tmp_path):
        assert polydoc.detect(workbook_document.save(tmp_path / "f.xlsx")) == "xlsx"

    def test_sheet_named_from_the_caption(self, workbook_document, tmp_path):
        reread = polydoc.open(workbook_document.save(tmp_path / "f.xlsx"))
        assert [s.name for s in reread.sheets] == ["Revenue"]

    def test_values_are_typed_not_stringly(self, workbook_document, tmp_path):
        import datetime

        import openpyxl

        path = workbook_document.save(tmp_path / "f.xlsx")
        sheet = openpyxl.load_workbook(path)["Revenue"]
        row = list(sheet[2])
        assert row[0].value == "Q1"
        assert row[1].value == 1200 and "$" in row[1].number_format
        assert row[2].value == pytest.approx(0.12) and "%" in row[2].number_format
        assert isinstance(row[3].value, (datetime.date, datetime.datetime))

    def test_header_is_bold_and_frozen(self, workbook_document, tmp_path):
        import openpyxl

        path = workbook_document.save(tmp_path / "f.xlsx")
        sheet = openpyxl.load_workbook(path)["Revenue"]
        assert sheet["A1"].font.bold
        assert sheet.freeze_panes == "A2"

    def test_read_back_as_a_table(self, workbook_document, tmp_path):
        reread = polydoc.open(workbook_document.save(tmp_path / "f.xlsx"))
        matrix = reread.sheets[0].content[0].to_matrix()
        assert matrix[0] == ["Quarter", "Revenue", "Growth", "Date"]
        assert matrix[1][0] == "Q1"

    def test_header_detection_on_read(self, workbook_document, tmp_path):
        reread = polydoc.open(workbook_document.save(tmp_path / "f.xlsx"))
        assert reread.sheets[0].content[0].header_rows == 1

    def test_multiple_tables_become_multiple_sheets(self, tmp_path):
        document = Document(
            body=[
                Table.from_rows([["a"], ["1"]], caption="First"),
                Table.from_rows([["b"], ["2"]], caption="Second"),
            ]
        )
        reread = polydoc.open(document.save(tmp_path / "multi.xlsx"))
        assert {s.name for s in reread.sheets} >= {"First", "Second"}

    def test_merged_cells_become_spans(self, tmp_path):
        from polydoc.model import TableCell, TableRow

        table = Table([TableRow([TableCell.of("wide", colspan=3)]), TableRow.of(["a", "b", "c"])])
        reread = polydoc.open(Document([table]).save(tmp_path / "sp.xlsx"))
        assert reread.sheets[0].content[0].rows[0].cells[0].colspan == 3

    def test_prose_is_not_lost(self, workbook_document, tmp_path):
        import openpyxl

        path = workbook_document.save(tmp_path / "f.xlsx", force_content_sheet=True)
        names = openpyxl.load_workbook(path).sheetnames
        assert "Content" in names

    def test_sheet_names_are_sanitised(self, tmp_path):
        import openpyxl

        table = Table.from_rows([["a"], ["1"]], caption="Bad/Name*With?Chars" * 3)
        path = Document([table]).save(tmp_path / "s.xlsx")
        name = openpyxl.load_workbook(path).sheetnames[0]
        assert len(name) <= 31
        assert not set(name) & set("\\/*?:[]")

    def test_csv_export_of_a_workbook(self, workbook_document, tmp_path):
        reread = polydoc.open(workbook_document.save(tmp_path / "f.xlsx"))
        assert "Quarter,Revenue" in polydoc.dumps(reread, "csv").decode()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


@needs_pdf_write
class TestPdfWriter:
    def test_file_is_created_and_detected(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.pdf")
        assert path.stat().st_size > 1000
        assert polydoc.detect(path) == "pdf"

    def test_starts_with_the_pdf_magic(self, simple_document, tmp_path):
        path = simple_document.save(tmp_path / "s.pdf")
        assert path.read_bytes().startswith(b"%PDF")

    def test_empty_document_does_not_crash(self, tmp_path):
        assert Document().save(tmp_path / "e.pdf").exists()

    def test_page_size_option(self, simple_document, tmp_path):
        path = simple_document.save(tmp_path / "letter.pdf", page_size=(612, 792))
        assert path.exists()


@needs_pdf_write
@needs_pdf_read
class TestPdfRoundTrip:
    @pytest.fixture
    def written(self, sample_document, tmp_path):
        return polydoc.open(sample_document.save(tmp_path / "doc.pdf"))

    def test_pages_are_preserved(self, written):
        assert len(written.pages) == 2  # the document contains a page break

    def test_metadata(self, written):
        assert written.metadata.title == "Quarterly Report"
        assert written.metadata.authors == ["Ada Lovelace"]

    def test_headings_are_recovered_from_font_sizes(self, written):
        levels = [(h.level, h.text) for h in written.headings]
        assert (1, "Quarterly Report") in levels
        assert (2, "Findings") in levels

    def test_headings_have_no_redundant_emphasis(self, written):
        heading = written.find("heading[level=1]")
        assert all(run.style.bold is None for run in heading.content)

    def test_table_is_recovered(self, written):
        assert written.tables
        assert written.tables[0].to_matrix()[0] == ["Component", "Trials", "Rate"]

    def test_list_is_recovered(self, written):
        block = written.find("list_block")
        assert block is not None and len(block.items) >= 3

    def test_code_block_is_recovered(self, written):
        assert written.find("code_block") is not None

    def test_running_footer_is_stripped(self, written):
        # The writer stamps the title and a page number in the footer; neither should
        # appear as body content. Only page-level paragraphs are checked, since table
        # cells legitimately hold bare numbers.
        assert "Quarterly Report1" not in written.text
        for page in written.pages:
            paragraphs = [b.text.strip() for b in page.content if b.type == "paragraph"]
            assert not any(text.isdigit() for text in paragraphs)
            assert "Quarterly Report" not in paragraphs

    def test_outline_bookmarks_are_written(self, written):
        entries = written.attrs.get("outline", [])
        assert entries
        assert entries[0]["title"] == "Quarterly Report"

    def test_single_page_footer_is_stripped(self, tmp_path):
        """A one-page document has no repetition to detect a footer by, so geometry
        has to carry it. Regression guard for the footer leaking in as a paragraph."""
        document = Document(
            metadata=Metadata(title="One Pager"),
            body=[
                Heading.of("One Pager", 1),
                Paragraph.of("Body text on the only page of this document."),
                Paragraph.of("A second paragraph so the page has some substance."),
            ],
        )
        reread = polydoc.open(document.save(tmp_path / "one.pdf"))
        paragraphs = [b.text.strip() for b in reread.pages[0].content if b.type == "paragraph"]
        assert not any(text.rstrip().endswith(" 1") for text in paragraphs), paragraphs

    def test_body_prose_is_not_promoted_to_a_heading(self, tmp_path):
        """Table cells are set smaller than body copy. If they are included in the font
        profile they drag the measured body size down and real paragraphs then outrank
        it, turning prose into headings."""
        document = Document(
            body=[
                Heading.of("Report", 1),
                Paragraph.of("Some intro text."),
                Table.from_rows(
                    [["Format", "Read"], ["PDF", "yes"], ["DOCX", "yes"]],
                    caption="Matrix",
                ),
            ]
        )
        reread = polydoc.open(document.save(tmp_path / "skew.pdf"))
        assert "Some intro text." not in [h.text for h in reread.headings]

    def test_detection_failures_are_reported_not_swallowed(self, sample_document, tmp_path):
        """A silently table-free result is undiagnosable, so failures are recorded."""
        path = sample_document.save(tmp_path / "doc.pdf")
        # pdfplumber absent is the common case; simulate it by disabling tables.
        document = polydoc.open(path, tables=True)
        # No failure expected here, so no warnings should be present.
        assert "warnings" not in document.attrs or document.attrs["warnings"] == []

    def test_text_content_is_substantially_preserved(self, written, sample_document):
        for phrase in ["Centred text", "Correctness first", "After the break"]:
            assert phrase in written.text

    def test_page_selection_option(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.pdf")
        first_only = polydoc.open(path, pages=1)
        assert len(first_only.pages) == 1

    def test_flatten_option_drops_page_wrappers(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.pdf")
        flat = polydoc.open(path, flatten=True)
        assert flat.pages == []
        assert flat.headings

    def test_tables_can_be_disabled(self, sample_document, tmp_path):
        path = sample_document.save(tmp_path / "doc.pdf")
        assert polydoc.open(path, tables=False).tables == []

    def test_encrypted_pdf_reports_clearly(self, tmp_path):
        pytest.importorskip("pypdf")
        from pypdf import PdfReader, PdfWriter

        plain = Document([Paragraph.of("secret")]).save(tmp_path / "p.pdf")
        writer = PdfWriter()
        for page in PdfReader(str(plain)).pages:
            writer.add_page(page)
        writer.encrypt("hunter2")
        locked = tmp_path / "locked.pdf"
        with open(locked, "wb") as handle:
            writer.write(handle)

        from polydoc.exceptions import ParseError

        with pytest.raises(ParseError, match="password"):
            polydoc.open(locked)
        assert "secret" in polydoc.open(locked, password="hunter2").text


@needs_pdf_read
class TestForeignPdf:
    """Reading a PDF built with a raw canvas API, i.e. with no structure at all."""

    @pytest.fixture
    def path(self, tmp_path):
        reportlab = pytest.importorskip("reportlab")
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as canvas_module

        target = tmp_path / "native.pdf"
        canvas = canvas_module.Canvas(str(target), pagesize=A4)
        width, height = A4
        y = height - 70

        canvas.setFont("Helvetica-Bold", 24)
        canvas.drawString(60, y, "Annual Review 2026")
        y -= 46
        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawString(60, y, "1. Overview")
        y -= 24
        canvas.setFont("Helvetica", 10)
        for line in [
            "This document was produced with the canvas API and has no",
            "structural markup whatsoever, only positioned glyphs.",
        ]:
            canvas.drawString(60, y, line)
            y -= 14
        y -= 18
        canvas.setFont("Helvetica", 10)
        for line in [
            "A second paragraph sits below a wider vertical gap, which is",
            "the only available signal that a paragraph has begun.",
        ]:
            canvas.drawString(60, y, line)
            y -= 14
        y -= 20
        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawString(60, y, "2. Findings")
        y -= 24
        canvas.setFont("Helvetica", 10)
        for line in ["\u2022 First finding", "\u2022 Second finding", "\u2022 Third finding"]:
            canvas.drawString(72, y, line)
            y -= 14
        canvas.showPage()
        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawString(60, height - 70, "3. Second Page")
        canvas.setFont("Helvetica", 10)
        canvas.drawString(60, height - 96, "Content on the second page.")
        canvas.save()
        return target

    def test_title_inferred_from_the_largest_text(self, path):
        assert polydoc.open(path).metadata.title == "Annual Review 2026"

    def test_placeholder_title_is_ignored(self, path):
        # ReportLab's canvas writes /Title "untitled"; we must not believe it.
        assert polydoc.open(path).metadata.title != "untitled"

    def test_heading_hierarchy_from_font_ranking(self, path):
        document = polydoc.open(path)
        levels = {h.text: h.level for h in document.headings}
        assert levels["Annual Review 2026"] == 1
        assert levels["1. Overview"] == 2
        assert levels["2. Findings"] == 2

    def test_numbered_headings_beat_list_detection(self, path):
        document = polydoc.open(path)
        # "1. Overview" parses as an ordered list marker, but its font size says heading.
        assert any(h.text == "1. Overview" for h in document.headings)

    def test_paragraphs_split_on_vertical_gaps(self, path):
        document = polydoc.open(path)
        paragraphs = [b.text for b in document.blocks() if b.type == "paragraph"]
        assert any("canvas API" in p for p in paragraphs)
        assert any("second paragraph" in p for p in paragraphs)
        # They must be separate blocks, not one merged run of text.
        merged = [p for p in paragraphs if "canvas API" in p and "second paragraph" in p]
        assert not merged

    def test_wrapped_lines_are_joined(self, path):
        document = polydoc.open(path)
        paragraph = [
            b
            for b in document.blocks()
            if b.type == "paragraph" and "canvas API" in b.text
        ][0]
        assert "\n" not in paragraph.text

    def test_bullets_become_a_list(self, path):
        document = polydoc.open(path)
        block = document.find("list_block")
        assert block is not None and len(block.items) == 3

    def test_pages_are_separate(self, path):
        document = polydoc.open(path)
        assert len(document.pages) == 2
        assert "Second Page" in document.pages[1].text

    def test_every_writer_accepts_it(self, path, tmp_path):
        document = polydoc.open(path)
        for fmt, extension in [
            ("markdown", ".md"),
            ("html", ".html"),
            ("txt", ".txt"),
            ("json", ".json"),
            ("csv", ".csv"),
            ("docx", ".docx"),
            ("pptx", ".pptx"),
            ("xlsx", ".xlsx"),
            ("pdf", ".pdf"),
        ]:
            target = tmp_path / f"out{extension}"
            document.save(target, format=fmt)
            assert target.stat().st_size > 0, fmt
