"""XLSX reader, built on openpyxl.

Each worksheet becomes a :class:`~polydoc.model.Container` with ``role="sheet"``
holding one :class:`~polydoc.model.Table`, so a workbook keeps its shape and
``document.sheets`` gives direct access.

Details that matter for fidelity:

* **Merged ranges** become ``colspan``/``rowspan`` on the anchor cell; openpyxl reports
  the other covered cells as ``None``, and they are skipped.
* **Number formats** are honoured when rendering to text, so a date cell reads as a date
  rather than a serial number, and currency keeps its shape.
* **Used range only.** ``max_row``/``max_column`` can be inflated by stale formatting, so
  trailing empty rows and columns are trimmed.
"""

from __future__ import annotations

import datetime as _datetime
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from ...model import (
    Alignment,
    Block,
    Container,
    Document,
    Paragraph,
    ParagraphStyle,
    Table,
    TableCell,
    TableRow,
    Text,
    TextStyle,
    VerticalAlign,
)
from ..base import Reader, require
from ..registry import register_reader
from ..source import Source

__all__ = ["XlsxReader"]

#: Excel's epoch for the 1900 date system, offset for its leap-year quirk.
_EXCEL_EPOCH = _datetime.datetime(1899, 12, 30)


@register_reader
class XlsxReader(Reader):
    """Reads Excel workbooks (``.xlsx``, ``.xlsm``)."""

    format = "xlsx"
    extensions = (".xlsx", ".xlsm", ".xltx")
    aliases = ("excel", "spreadsheet")
    mime_types = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",)
    extra = "xlsx"
    archive_based = True
    description = "Microsoft Excel (Open XML), one table per worksheet"

    def read(self, source: Source, **options: Any) -> Document:
        self.enforce_limits(source, **options)
        openpyxl = require("openpyxl", "Reading XLSX", extra="xlsx")

        data_only: bool = options.get("data_only", True)
        include_hidden: bool = options.get("include_hidden", False)
        header_row: Optional[bool] = options.get("header_row")
        sheet_filter = options.get("sheets")
        keep_styles: bool = options.get("keep_styles", True)

        workbook = openpyxl.load_workbook(
            source.stream(),
            data_only=data_only,
            read_only=False,
            keep_links=False,
        )

        document = Document()
        try:
            for worksheet in workbook.worksheets:
                if not include_hidden and worksheet.sheet_state != "visible":
                    continue
                if sheet_filter and worksheet.title not in sheet_filter:
                    continue
                table = self._sheet_table(worksheet, header_row, keep_styles)
                if table is None:
                    continue
                container = Container(
                    [table], role="sheet", name=worksheet.title
                )
                container.attrs["sheet_state"] = worksheet.sheet_state
                document.append(container)

            self._read_metadata(workbook, document)
        finally:
            workbook.close()

        if document.metadata.title is None and document.sheets:
            document.metadata.title = document.sheets[0].name

        return self.finalise(document, source)

    # -- sheets ---------------------------------------------------------------
    def _sheet_table(
        self,
        worksheet: Any,
        header_row: Optional[bool],
        keep_styles: bool,
    ) -> Optional[Table]:
        bounds = self._used_bounds(worksheet)
        if bounds is None:
            return None
        max_row, max_column = bounds

        merges = self._merge_map(worksheet)
        covered = merges["covered"]
        spans: Dict[Tuple[int, int], Tuple[int, int]] = merges["spans"]

        rows: List[TableRow] = []
        for row_index in range(1, max_row + 1):
            cells: List[TableCell] = []
            for column_index in range(1, max_column + 1):
                if (row_index, column_index) in covered:
                    continue
                native = worksheet.cell(row=row_index, column=column_index)
                colspan, rowspan = spans.get((row_index, column_index), (1, 1))
                cells.append(self._cell(native, colspan, rowspan, keep_styles))
            if cells:
                rows.append(TableRow(cells))

        if not rows:
            return None

        use_header = header_row if header_row is not None else self._looks_like_header(rows)
        if use_header:
            rows[0].is_header = True

        table = Table(rows, header_rows=1 if use_header else 0)
        widths = self._column_widths(worksheet, max_column)
        if widths:
            table.column_widths = widths
        if getattr(worksheet, "freeze_panes", None):
            table.attrs["freeze_panes"] = worksheet.freeze_panes
        return table

    @staticmethod
    def _used_bounds(worksheet: Any) -> Optional[Tuple[int, int]]:
        """Trim the reported dimensions to rows and columns that hold values.

        Excel files routinely report a max_row in the thousands because of stale
        formatting; building a table that size would be both slow and wrong.
        """
        max_row = worksheet.max_row or 0
        max_column = worksheet.max_column or 0
        if not max_row or not max_column:
            return None

        last_row = 0
        last_column = 0
        for row in worksheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_column
        ):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    if cell.row > last_row:
                        last_row = cell.row
                    if cell.column > last_column:
                        last_column = cell.column
        if not last_row or not last_column:
            return None
        return (last_row, last_column)

    @staticmethod
    def _merge_map(worksheet: Any) -> Dict[str, Any]:
        """Map merged ranges to spans, and note which cells they swallow."""
        covered: Set[Tuple[int, int]] = set()
        spans: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for merged in getattr(worksheet, "merged_cells", {}).ranges:
            min_row, min_col = merged.min_row, merged.min_col
            max_row, max_col = merged.max_row, merged.max_col
            spans[(min_row, min_col)] = (
                max_col - min_col + 1,
                max_row - min_row + 1,
            )
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    if (row, column) != (min_row, min_col):
                        covered.add((row, column))
        return {"covered": covered, "spans": spans}

    def _cell(
        self,
        native: Any,
        colspan: int,
        rowspan: int,
        keep_styles: bool,
    ) -> TableCell:
        text = self._format_value(native)
        style = TextStyle()
        para_style = ParagraphStyle()
        background = None
        valign = None

        if keep_styles:
            style = self._text_style(native)
            para_style, valign = self._alignment(native)
            background = self._fill(native)

        content: List[Block] = []
        if text:
            content.append(Paragraph([Text(text, style)], para_style))

        cell = TableCell(
            content,
            colspan=colspan,
            rowspan=rowspan,
            valign=valign,
            background=background,
        )
        if native.data_type == "f" and isinstance(native.value, str):
            cell.attrs["formula"] = native.value
        if getattr(native, "hyperlink", None) is not None:
            target = getattr(native.hyperlink, "target", None)
            if target:
                cell.attrs["hyperlink"] = target
        return cell

    @staticmethod
    def _format_value(native: Any) -> str:
        """Render a cell value as text, respecting its number format."""
        value = native.value
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (_datetime.datetime, _datetime.date, _datetime.time)):
            if isinstance(value, _datetime.datetime):
                if value.time() == _datetime.time(0, 0):
                    return value.date().isoformat()
                return value.isoformat(sep=" ")
            return value.isoformat()

        number_format = (getattr(native, "number_format", "") or "").lower()
        if isinstance(value, (int, float)):
            if "%" in number_format:
                return f"{value * 100:g}%"
            if isinstance(value, float) and value.is_integer() and "0.0" not in number_format:
                return str(int(value))
            return f"{value:g}" if isinstance(value, float) else str(value)
        return str(value)

    @staticmethod
    def _text_style(native: Any) -> TextStyle:
        font = getattr(native, "font", None)
        if font is None:
            return TextStyle()
        colour = None
        if font.color is not None and getattr(font.color, "rgb", None):
            rgb = str(font.color.rgb)
            # openpyxl reports ARGB; drop the alpha channel.
            if len(rgb) == 8:
                rgb = rgb[2:]
            if len(rgb) == 6 and rgb.upper() != "000000":
                colour = f"#{rgb.lower()}"
        return TextStyle(
            bold=True if font.bold else None,
            italic=True if font.italic else None,
            underline=True if font.underline else None,
            strike=True if font.strike else None,
            font_family=font.name or None,
            font_size=float(font.size) if font.size else None,
            color=colour,
        )

    @staticmethod
    def _alignment(native: Any) -> Tuple[ParagraphStyle, Optional[VerticalAlign]]:
        alignment = getattr(native, "alignment", None)
        if alignment is None:
            return ParagraphStyle(), None
        return (
            ParagraphStyle(alignment=Alignment.coerce(alignment.horizontal)),
            VerticalAlign.coerce(alignment.vertical),
        )

    @staticmethod
    def _fill(native: Any) -> Optional[str]:
        fill = getattr(native, "fill", None)
        if fill is None or getattr(fill, "fill_type", None) not in ("solid",):
            return None
        colour = getattr(fill, "start_color", None) or getattr(fill, "fgColor", None)
        rgb = getattr(colour, "rgb", None)
        if not rgb:
            return None
        rgb = str(rgb)
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) != 6 or rgb.upper() in ("FFFFFF", "000000"):
            return None
        return f"#{rgb.lower()}"

    @staticmethod
    def _looks_like_header(rows: Sequence[TableRow]) -> bool:
        """Treat row 1 as a header when it is textual, bold, or unlike row 2."""
        if len(rows) < 2:
            return False
        first, second = rows[0], rows[1]
        first_texts = [cell.text.strip() for cell in first.cells]
        if not any(first_texts):
            return False

        bold = sum(
            1
            for cell in first.cells
            for block in cell.content
            if isinstance(block, Paragraph)
            for run in block.content
            if isinstance(run, Text) and run.style.bold
        )
        if bold and bold >= len([t for t in first_texts if t]) / 2:
            return True

        def numeric(value: str) -> bool:
            try:
                float(value.replace(",", "").replace("%", "").replace("$", ""))
                return True
            except ValueError:
                return False

        first_numeric = sum(1 for t in first_texts if t and numeric(t))
        second_texts = [cell.text.strip() for cell in second.cells]
        second_numeric = sum(1 for t in second_texts if t and numeric(t))
        # A textual first row above a numeric second row is a header.
        return first_numeric == 0 and second_numeric > 0

    @staticmethod
    def _column_widths(worksheet: Any, max_column: int) -> Optional[List[float]]:
        dimensions = getattr(worksheet, "column_dimensions", None)
        if not dimensions:
            return None
        from openpyxl.utils import get_column_letter

        widths: List[float] = []
        for index in range(1, max_column + 1):
            letter = get_column_letter(index)
            dimension = dimensions.get(letter)
            width = getattr(dimension, "width", None) if dimension is not None else None
            # Excel widths are in character units; ~7 points per character.
            widths.append(round(float(width) * 7.0, 2) if width else 0.0)
        return widths if any(widths) else None

    @staticmethod
    def _read_metadata(workbook: Any, document: Document) -> None:
        props = getattr(workbook, "properties", None)
        if props is None:
            return
        meta = document.metadata
        meta.title = props.title or None
        if props.creator and props.creator != "openpyxl":
            meta.authors = [p.strip() for p in props.creator.split(";") if p.strip()]
        meta.subject = props.subject or None
        if props.keywords:
            import re

            meta.keywords = [p.strip() for p in re.split(r"[;,]", props.keywords) if p.strip()]
        meta.description = props.description or None
        meta.category = props.category or None
        meta.language = props.language or None
        meta.created = props.created
        meta.modified = props.modified
