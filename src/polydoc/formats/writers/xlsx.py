"""XLSX writer, built on openpyxl.

A spreadsheet is a grid, so non-tabular content has to go somewhere sensible rather
than be dropped. The strategy:

* Each ``role="sheet"`` container, or each table, becomes its own worksheet.
* Prose surrounding tables is written to a leading "Content" sheet as ``type``/``text``
  rows, so a converted report keeps its narrative alongside its data.
* Values are *typed* on the way in: a cell reading "1,234" becomes the number 1234 and
  "2024-03-01" becomes a date, so the result is usable for calculation rather than a
  grid of strings.
"""

from __future__ import annotations

import datetime as _datetime
import re
from typing import Any, BinaryIO, Dict, List, Optional, Sequence, Tuple

from ...model import (
    Alignment,
    Block,
    BlockContainer,
    Container,
    Document,
    Heading,
    Page,
    Paragraph,
    Section,
    Slide,
    Table,
    TableCell,
    Text,
    VerticalAlign,
)
from ..base import Writer, require
from ..registry import register_writer

__all__ = ["XlsxWriter"]

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_MAX_SHEET_NAME = 31

_NUMBER_RE = re.compile(r"^[-+]?(\d{1,3}(,\d{3})*|\d+)(\.\d+)?$")
_PERCENT_RE = re.compile(r"^[-+]?\d+(\.\d+)?\s*%$")
_CURRENCY_RE = re.compile(r"^([$\u00a3\u20ac\u00a5])\s*[-+]?(\d{1,3}(,\d{3})*|\d+)(\.\d+)?$")
_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y")

_HORIZONTAL = {
    Alignment.LEFT: "left",
    Alignment.CENTER: "center",
    Alignment.RIGHT: "right",
    Alignment.JUSTIFY: "justify",
}
_VERTICAL = {
    VerticalAlign.TOP: "top",
    VerticalAlign.MIDDLE: "center",
    VerticalAlign.BOTTOM: "bottom",
}


@register_writer
class XlsxWriter(Writer):
    """Writes Excel workbooks (``.xlsx``)."""

    format = "xlsx"
    extensions = (".xlsx",)
    aliases = ("excel", "spreadsheet")
    mime_types = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",)
    extra = "xlsx"
    description = "Microsoft Excel (Open XML), one worksheet per table"

    def write(self, document: Document, stream: BinaryIO, **options: Any) -> None:
        openpyxl = require("openpyxl", "Writing XLSX", extra="xlsx")
        builder = _XlsxBuilder(openpyxl, document, options)
        builder.build().save(stream)


class _XlsxBuilder:
    def __init__(self, openpyxl: Any, document: Document, options: Dict[str, Any]) -> None:
        self.openpyxl = openpyxl
        self.document = document
        self.options = options
        self.coerce_types: bool = options.get("coerce_types", True)
        self.autofit: bool = options.get("autofit", True)
        self.freeze_header: bool = options.get("freeze_header", True)
        self.include_prose: bool = options.get("include_prose", True)
        self.workbook = openpyxl.Workbook()
        self._used_names: List[str] = []

    # -- entry point ----------------------------------------------------------
    def build(self) -> Any:
        groups = self._collect()
        default = self.workbook.active
        created = False

        prose = self._prose_rows() if self.include_prose else []
        if prose and (len(groups) != 1 or self.options.get("force_content_sheet")):
            sheet = default
            sheet.title = self._unique_name("Content")
            self._write_rows(sheet, prose, header=True)
            created = True

        for name, table in groups:
            if created:
                sheet = self.workbook.create_sheet(self._unique_name(name))
            else:
                sheet = default
                sheet.title = self._unique_name(name)
                created = True
            self._write_table(sheet, table)

        if not created:
            sheet = default
            sheet.title = self._unique_name(self.document.metadata.title or "Sheet1")
            rows = prose or [["type", "text"]]
            self._write_rows(sheet, rows, header=True)

        self._apply_metadata()
        return self.workbook

    def _apply_metadata(self) -> None:
        meta = self.document.metadata
        props = self.workbook.properties
        if meta.title:
            props.title = meta.title
        if meta.authors:
            props.creator = "; ".join(meta.authors)
        if meta.subject:
            props.subject = meta.subject
        if meta.keywords:
            props.keywords = ", ".join(meta.keywords)
        if meta.description:
            props.description = meta.description
        if meta.category:
            props.category = meta.category
        if meta.created:
            props.created = meta.created
        if meta.modified:
            props.modified = meta.modified

    # -- collection -----------------------------------------------------------
    def _collect(self) -> List[Tuple[str, Table]]:
        """Find every table, paired with the best available sheet name."""
        out: List[Tuple[str, Table]] = []

        def walk(blocks: Sequence[Block], label: Optional[str]) -> None:
            for block in blocks:
                if isinstance(block, Table):
                    name = block.caption or label or f"Table{len(out) + 1}"
                    out.append((name, block))
                elif isinstance(block, Container):
                    walk(block.content, block.name or label)
                elif isinstance(block, Slide):
                    walk(block.content, block.title or label)
                elif isinstance(block, Section):
                    walk(block.content, block.title_text or label)
                elif isinstance(block, Page):
                    page_label = f"Page {block.number}" if block.number else label
                    walk(block.content, page_label)
                elif isinstance(block, BlockContainer):
                    walk(block.content, label)

        walk(self.document.body, None)
        return out

    def _prose_rows(self) -> List[List[Any]]:
        """Flatten non-tabular content into ``type, text`` rows."""
        rows: List[List[Any]] = [["type", "level", "text"]]
        for block in self.document.blocks():
            if isinstance(block, (Table, TableCell)) or isinstance(block, BlockContainer):
                continue
            if any(isinstance(a, Table) for a in block.ancestors()):
                continue
            text = " ".join(block.text.split())
            if not text:
                continue
            level = block.level if isinstance(block, Heading) else ""
            rows.append([block.type, level, text])
        return rows if len(rows) > 1 else []

    def _unique_name(self, name: str) -> str:
        """Excel sheet names are limited to 31 chars and must be unique."""
        cleaned = _INVALID_SHEET_CHARS.sub("-", " ".join(str(name).split())).strip("'")
        cleaned = cleaned[:_MAX_SHEET_NAME] or "Sheet"
        candidate = cleaned
        counter = 2
        while candidate.lower() in (n.lower() for n in self._used_names):
            suffix = f" ({counter})"
            candidate = cleaned[: _MAX_SHEET_NAME - len(suffix)] + suffix
            counter += 1
        self._used_names.append(candidate)
        return candidate

    # -- writing --------------------------------------------------------------
    def _write_rows(self, sheet: Any, rows: Sequence[Sequence[Any]], header: bool) -> None:
        from openpyxl.styles import Font

        for row in rows:
            sheet.append(list(row))
        if header and rows:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            if self.freeze_header:
                sheet.freeze_panes = "A2"
        if self.autofit:
            self._autofit(sheet)

    def _write_table(self, sheet: Any, table: Table) -> None:
        from openpyxl.styles import Alignment as XlAlignment
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        row_cursor = 1
        # Reserve the grid positions consumed by row spans.
        occupied: Dict[Tuple[int, int], bool] = {}

        for row in table.rows:
            column_cursor = 1
            for cell in row.cells:
                while occupied.get((row_cursor, column_cursor)):
                    column_cursor += 1

                target = sheet.cell(row=row_cursor, column=column_cursor)
                value, number_format = self._cell_value(cell)
                target.value = value
                if number_format:
                    target.number_format = number_format

                style = self._cell_text_style(cell)
                if style is not None or row.is_header:
                    target.font = Font(
                        bold=bool(row.is_header or (style and style.bold)),
                        italic=bool(style and style.italic),
                        underline="single" if style and style.underline else None,
                        strike=bool(style and style.strike),
                        name=(style.font_family if style and style.font_family else None),
                        size=(style.font_size if style and style.font_size else None),
                        color=(
                            style.color.lstrip("#").upper()
                            if style and style.color
                            else None
                        ),
                    )

                alignment = self._cell_alignment(cell)
                if alignment is not None:
                    target.alignment = XlAlignment(**alignment)
                if cell.background:
                    fill = cell.background.lstrip("#").upper()
                    target.fill = PatternFill("solid", start_color=fill, end_color=fill)

                if cell.attrs.get("hyperlink"):
                    target.hyperlink = cell.attrs["hyperlink"]

                if cell.colspan > 1 or cell.rowspan > 1:
                    sheet.merge_cells(
                        start_row=row_cursor,
                        start_column=column_cursor,
                        end_row=row_cursor + cell.rowspan - 1,
                        end_column=column_cursor + cell.colspan - 1,
                    )
                for r in range(row_cursor, row_cursor + cell.rowspan):
                    for c in range(column_cursor, column_cursor + cell.colspan):
                        occupied[(r, c)] = True

                column_cursor += cell.colspan
            row_cursor += 1

        if table.header_rows and self.freeze_header:
            sheet.freeze_panes = f"A{table.header_rows + 1}"
        elif table.attrs.get("freeze_panes"):
            sheet.freeze_panes = table.attrs["freeze_panes"]

        if table.column_widths:
            total = sum(w for w in table.column_widths if w > 0)
            if total > 0:
                for index, width in enumerate(table.column_widths, start=1):
                    if width > 0:
                        sheet.column_dimensions[get_column_letter(index)].width = max(
                            6.0, round(width / 7.0, 1)
                        )
        elif self.autofit:
            self._autofit(sheet)

    def _autofit(self, sheet: Any) -> None:
        """Approximate column auto-width from the longest rendered value."""
        from openpyxl.utils import get_column_letter

        widths: Dict[int, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = max(len(line) for line in str(cell.value).split("\n"))
                widths[cell.column] = max(widths.get(cell.column, 0), length)
        for column, length in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = min(
                60.0, max(8.0, length + 2)
            )

    # -- cell helpers ---------------------------------------------------------
    def _cell_value(self, cell: TableCell) -> Tuple[Any, Optional[str]]:
        if cell.attrs.get("formula"):
            return (cell.attrs["formula"], None)
        text = cell.text.strip()
        if not text:
            return (None, None)
        if not self.coerce_types:
            return (text, None)
        return _coerce(text)

    @staticmethod
    def _cell_text_style(cell: TableCell) -> Optional[Any]:
        for block in cell.content:
            if isinstance(block, Paragraph):
                for run in block.content:
                    if isinstance(run, Text) and not run.style.is_empty():
                        return run.style
        return None

    @staticmethod
    def _cell_alignment(cell: TableCell) -> Optional[Dict[str, Any]]:
        options: Dict[str, Any] = {}
        for block in cell.content:
            align = getattr(getattr(block, "style", None), "alignment", None)
            if align is not None:
                horizontal = _HORIZONTAL.get(align)
                if horizontal:
                    options["horizontal"] = horizontal
                break
        if cell.valign is not None:
            vertical = _VERTICAL.get(cell.valign)
            if vertical:
                options["vertical"] = vertical
        if "\n" in cell.text:
            options["wrap_text"] = True
        return options or None


def _coerce(text: str) -> Tuple[Any, Optional[str]]:
    """Convert display text into a typed value plus a number format.

    Keeping numbers as numbers is the difference between a spreadsheet you can compute
    with and a grid of strings.

    >>> _coerce("1,234")
    (1234, '#,##0')
    >>> _coerce("45%")[0]
    0.45
    """
    if text.upper() in ("TRUE", "FALSE"):
        return (text.upper() == "TRUE", None)

    percent = _PERCENT_RE.match(text)
    if percent:
        try:
            return (float(text.rstrip("% ").replace(",", "")) / 100.0, "0.0%")
        except ValueError:
            pass

    currency = _CURRENCY_RE.match(text)
    if currency:
        symbol = currency.group(1)
        try:
            value = float(re.sub(r"[^\d.\-+]", "", text))
            return (value, f'"{symbol}"#,##0.00')
        except ValueError:
            pass

    if _NUMBER_RE.match(text):
        cleaned = text.replace(",", "")
        try:
            if "." in cleaned:
                return (float(cleaned), "#,##0.00" if "," in text else None)
            return (int(cleaned), "#,##0" if "," in text else None)
        except ValueError:
            pass

    for fmt in _DATE_FORMATS:
        try:
            parsed = _datetime.datetime.strptime(text, fmt)
            return (parsed.date(), "yyyy-mm-dd")
        except ValueError:
            continue

    # An overlong value would break Excel's 32767-character cell limit.
    return (text[:32767], None)
